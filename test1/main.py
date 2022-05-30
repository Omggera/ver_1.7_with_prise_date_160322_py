from tkinter import *
from tkinter import filedialog
from tkinter.ttk import Combobox
from tkinter import ttk
import openpyxl
from openpyxl.styles import Border, Side, Alignment
from openpyxl.drawing.image import Image
import fonts
import borders
from datetime import datetime, timedelta
import re
import warnings


window = Tk()
window.title('Ленивый набор')
window.geometry('550x480')
window['bg'] = 'lightsteelblue1'

lsb1 = 'lightsteelblue1'
sb3 = 'steelblue3'
style = ttk.Style()
style.theme_create('theme1',
                   parent='alt',
                   settings={'TNotebook': {'configure': {'tabmargins': [2, 5, 2, 0] } },
                             'TNotebook.Tab': {'configure': {'padding': [5, 1], 'background': lsb1},
                                               'map':       {'background': [('selected', sb3)],
                                                             'foreground': [('selected', 'white')],
                                                             'expand': [('selected', [1, 1, 1, 0])]
                                                             }
                                               }
                             }
                   )

style.theme_use("theme1")
ttk.Style().configure('TNotebook', background='lightsteelblue1')
ttk.Style().configure('TFrame', background='lightsteelblue1')

tab = ttk.Notebook(window)
tab1 = ttk.Frame(tab)
tab.add(tab1, text='Заказы')

tab2 = ttk.Frame(tab)
tab.add(tab2, text='Прайс')
tab.grid(column=0, row=0, sticky=W)


def clicked():
    file_open = filedialog.askopenfilenames()
    names = file_open
    for pp in names:
        lb.insert(END, pp)


def savefold():
    dire = filedialog.askdirectory()
    names = dire
    for nn in names:
        lbl.configure(text=f'{names}')


lbl2 = Label(tab1, text='Выберите город', font=('Arial Bold', 10))
lbl2.grid(column=2, row=1, padx=3, pady=8)

lbl = Label(tab1,
            text='Нажмите кнопку и выберите папку для сохранения            ',
            font=('Arial Bold', 10),
            bg='white',
            height=1,
            width=44)
lbl.grid(column=0, row=1, sticky=W, columnspan=1, padx=3, pady=8)
btn = Button(tab1, text='...', command=savefold)
btn.grid(column=1, row=1, sticky=W, pady=8)

lb = Listbox(tab1, height=25, width=60, selectmode=EXTENDED)
lb.grid(column=0, row=2, rowspan=6, padx=3)
scroll = Scrollbar(tab1, command=lb.yview)
scroll.grid(column=1, row=2, columnspan=1, rowspan=6, sticky=NS)
lb.config(yscrollcommand=scroll.set)

combo = Combobox(tab1, font=('Arial Bold', 10), width=15)
combo['values'] = ('Владимир', 'Иваново', 'Муром', 'Ковров')
combo['state'] = 'readonly'
combo.current()
combo.grid(column=2, row=2, padx=5, sticky=N)


def city_changed(eventObject):
    eventObject = eventObject
    eventObject = combo.get()
    #print(eventObject)
    return eventObject
combo.bind('<<ComboboxSelected>>', city_changed)


'''bar = Progressbar(window, orient=HORIZONTAL, length=360, mode='determinate')
bar.grid(column=0, row=7, padx=3, pady=5, columnspan=1, sticky=N)
txt = Label(window, text='0%')
txt.grid(column=1, row=7, columnspan=2, sticky=W)'''


def delete_all():
    lb.delete(0, 'end')


def delete():
    selection = lb.curselection()
    lb.delete(selection[0])


def select():
    lb.select_set(0, END)


def nabor():
    r = lb.get(0, 70)
    for i in range(0, len(r), 1):
        file = r[i:i + 1]
        files = ''.join(file)
        wb1 = openpyxl.load_workbook(files)
        ws1 = wb1.active
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active

        # Настройка масштабирования и полей
        ws2.sheet_view.zoomScale = 115
        ws2.page_setup.scale = 78
        ws2.page_setup.paperSize = 4
        ws2.page_margins.left = 0.18
        ws2.page_margins.right = 0.05
        ws2.page_margins.top = 0.10
        ws2.page_margins.bottom = 0.05
        ws2.page_margins.header = 0.10
        ws2.page_margins.footer = 0.05

        # Задаю высоту и ширину для нужных ячеек
        ws2.row_dimensions[1].height = 12
        ws2.column_dimensions['A'].width = 4.86
        ws2.row_dimensions[2].height = 9.75
        ws2.column_dimensions['B'].width = 2.43 + 0.71
        ws2.row_dimensions[3].height = 10.5
        ws2.column_dimensions['C'].width = 5.71 + 0.71
        ws2.row_dimensions[4].height = 9
        ws2.column_dimensions['D'].width = 3.14 + 0.71
        ws2.row_dimensions[5].height = 10.5
        ws2.column_dimensions['E'].width = 2.43 + 0.71
        ws2.row_dimensions[6].height = 10.5
        ws2.column_dimensions['F'].width = 1.14 + 0.71
        ws2.row_dimensions[7].height = 10.5
        ws2.column_dimensions['G'].width = 0.75 + 0.71
        ws2.row_dimensions[8].height = 13.5
        ws2.column_dimensions['H'].width = 3.71 + 0.71
        ws2.row_dimensions[9].height = 3
        ws2.column_dimensions['I'].width = 3 + 0.71
        ws2.row_dimensions[10].height = 14
        ws2.column_dimensions['J'].width = 0.17 + 0.71
        ws2.row_dimensions[11].height = 3
        ws2.column_dimensions['K'].width = 24.57 + 0.71
        ws2.row_dimensions[12].height = 12
        ws2.column_dimensions['L'].width = 4.71 + 0.71
        ws2.row_dimensions[13].height = 2.25
        ws2.column_dimensions['M'].width = 5.14 + 0.71
        ws2.row_dimensions[14].height = 12
        ws2.column_dimensions['N'].width = 5.86 + 0.71
        ws2.row_dimensions[15].height = 5.25
        ws2.column_dimensions['O'].width = 2.57 + 0.71
        ws2.row_dimensions[16].height = 11.5
        ws2.column_dimensions['P'].width = 1.43 + 0.71
        ws2.row_dimensions[17].height = 11.5
        ws2.column_dimensions['Q'].width = 0.75 + 0.71
        ws2.row_dimensions[18].height = 11.5
        ws2.column_dimensions['R'].width = 2.43 + 0.71
        ws2.row_dimensions[19].height = 11.5
        ws2.column_dimensions['S'].width = 0.17 + 0.71
        ws2.row_dimensions[20].height = 11.5
        ws2.column_dimensions['T'].width = 4 + 0.71
        ws2.row_dimensions[21].height = 11.5
        ws2.column_dimensions['U'].width = 2.29 + 0.71
        ws2.row_dimensions[22].height = 11.5
        ws2.column_dimensions['V'].width = 0.33 + 0.71
        ws2.row_dimensions[23].height = 11.5
        ws2.column_dimensions['W'].width = 2.43 + 0.71
        ws2.row_dimensions[24].height = 11.5
        ws2.column_dimensions['X'].width = 3.43 + 0.71
        ws2.row_dimensions[25].height = 11.5
        ws2.column_dimensions['Y'].width = 0.92 + 0.71
        ws2.row_dimensions[26].height = 11.5
        ws2.column_dimensions['Z'].width = 20.57 + 0.71
        ws2.row_dimensions[27].height = 11.5
        ws2.column_dimensions['AA'].width = 0.02
        ws2.row_dimensions[28].height = 11.5
        ws2.column_dimensions['AB'].width = 0.08
        ws2.row_dimensions[29].height = 11.5
        ws2.row_dimensions[30].height = 11.5
        ws2.row_dimensions[31].height = 11.5
        ws2.row_dimensions[32].height = 2.25
        ws2.row_dimensions[33].height = 10.5
        ws2.row_dimensions[34].height = 8
        ws2.row_dimensions[35].height = 8
        ws2.row_dimensions[36].height = 8
        ws2.row_dimensions[37].height = 8
        ws2.row_dimensions[38].height = 8
        ws2.row_dimensions[39].height = 8
        ws2.row_dimensions[40].height = 8
        ws2.row_dimensions[41].height = 8
        ws2.row_dimensions[42].height = 8
        ws2.row_dimensions[43].height = 6
        ws2.row_dimensions[44].height = 11.25
        ws2.row_dimensions[45].height = 3.75
        ws2.row_dimensions[46].height = 5.25
        ws2.row_dimensions[47].height = 11.25
        ws2.row_dimensions[48].height = 9.75
        ws2.row_dimensions[49].height = 12.75
        ws2.row_dimensions[50].height = 10.5
        ws2.row_dimensions[51].height = 4
        ws2.row_dimensions[52].height = 20
        ws2.row_dimensions[53].height = 9.75
        ws2.row_dimensions[54].height = 9

        # Заполнение ячеек
        ws2.cell(row=4, column=6).value = ws1['H4'].value + f' г. {city_changed(combo.get())}'
        ws2.merge_cells(start_row=4, start_column=6, end_row=5, end_column=24)
        currentCell = ws2.cell(row=4, column=6)
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        ws2['F4'].font = fonts.font1

        city = ws2['F4'].value
        city_res = re.findall(r'\w+', city)
        city_word = city_res[10]

        ws2.merge_cells(start_row=2, start_column=9, end_row=3, end_column=20)
        currentCell = ws2.cell(row=2, column=9)
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        ws2['I2'].font = fonts.font1

        ws2['Z3'] = '8 (800) 555-87-78'
        ws2.merge_cells(start_row=3, start_column=26, end_row=3, end_column=27)
        currentCell = ws2.cell(row=3, column=26)
        currentCell.alignment = Alignment(horizontal='right', vertical='center')
        ws2['Z3'].font = fonts.font2

        ws2.merge_cells(start_row=6, start_column=26, end_row=6, end_column=27)
        currentCell = ws2.cell(row=6, column=26)
        currentCell.alignment = Alignment(horizontal='right', vertical='center')
        ws2['Z6'].font = fonts.font2

        ws2['X49'] = 'Трофимова Татьяна Павловна'
        if city_word == 'Владимир':
            shapka = 'ИП Трофимова Т.П. - TechPort.Vladimir'
            ws2['I2'] = f'  {shapka}'
            ws2['AK2'] = f'  {shapka}'
            ws2['I56'] = f'  {shapka}'
            ws2['Z60'] = '8 (4922) 46-20-57'
            ws2['Z6'] = '8 (4922) 46-20-57'
            ws2['BC6'] = '8 (4922) 46-20-57'
        elif city_word == 'Иваново':
            shapka = 'ИП Данилов С.Ю. - TechPort.Ivanovo'
            ws2['I2'] = f'  {shapka}'
            ws2['AK2'] = f'  {shapka}'
            ws2['I56'] = f'  {shapka}'
            ws2['Z60'] = '8 (9644) 96-02-01'
            ws2['Z6'] = '8 (9644) 96-02-01'
            ws2['BC6'] = '8 (9644) 96-02-01'
            ws2['X49'] = 'Данилов Степан Юрьевич'
        elif city_word == 'Муром':
            shapka = 'ИП Трофимова Т.П. - TechPort.Murom'
            ws2['I2'] = f'  {shapka}'
            ws2['AK2'] = f'  {shapka}'
            ws2['I56'] = f'  {shapka}'
            ws2['Z60'] = '8 (903) 833-55-15'
            ws2['Z6'] = '8 (903) 833-55-15'
            ws2['BC6'] = '8 (903) 833-55-15'
        elif city_word == 'Ковров':
            shapka = 'ИП Трофимова Т.П. - TechPort.Kovrov'
            ws2['I2'] = f'  {shapka}'
            ws2['AK2'] = f'  {shapka}'
            ws2['I56'] = f'  {shapka}'
            ws2['Z60'] = '8 (920) 949-04-30'
            ws2['Z6'] = '8 (920) 949-04-30'
            ws2['BC6'] = '8 (920) 949-04-30'

        ws2['Z2'] = 'Отдел продаж'
        ws2.merge_cells(start_row=2, start_column=26, end_row=2, end_column=27)
        currentCell = ws2.cell(row=2, column=26)
        currentCell.alignment = Alignment(horizontal='right', vertical='center')
        ws2['Z2'].font = fonts.font2

        ws2['Z5'] = 'Служба доставки'
        ws2.merge_cells(start_row=5, start_column=26, end_row=5, end_column=27)
        currentCell = ws2.cell(row=5, column=26)
        currentCell.alignment = Alignment(horizontal='right', vertical='center')
        ws2['Z5'].font = fonts.font2

        for row in ws2.iter_cols(min_col=2, max_col=28, min_row=7, max_row=14):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='bottom')

        ws2['B7'] = '7. ИНФОРМАЦИЯ О ТОВАРЕ И АДРЕСЕ ДОСТАВКИ'
        ws2.merge_cells(start_row=7, start_column=2, end_row=7, end_column=28)
        ws2['B7'].font = fonts.font3

        ws2['B8'] = 'Адрес:'
        ws2.merge_cells(start_row=8, start_column=2, end_row=8, end_column=3)
        ws2['B8'].font = fonts.font4

        # Адрес доставки
        ws2.merge_cells('F8:AB8')
        ws2['F8'].font = fonts.font1
        borders.set_border(ws2, 'F8:AB8')
        if city_word == 'Владимир':
            ws2['F8'] = 'г. Владимир'
            if ws1['F8'].value == None:
                ws2['F8'] = 'г. Владимир'
                if ws1['F8'].value == None:
                    ws2['F8'] = 'г. Владимир'
                else:
                    ws2.cell(row=8, column=6).value = ws1['F8'].value
            else:
                if len(ws1['F8'].value) == 0:
                    ws2['F8'] = 'г. Владимир'
                else:
                    ws2.cell(row=8, column=6).value = ws1['F8'].value

        elif city_word == 'Иваново':
            ws2['F8'] = 'г. Иваново'
            if ws1['F8'].value == None:
                ws2['F8'] = 'г. Иваново'
                if ws1['F8'].value == None:
                    ws2['F8'] = 'г. Иваново'
                else:
                    ws2.cell(row=8, column=6).value = ws1['F8'].value
            else:
                if len(ws1['F8'].value) == 0:
                    ws2['F8'] = 'г. Иваново'
                else:
                    ws2.cell(row=8, column=6).value = ws1['F8'].value

        elif city_word == 'Муром':
            ws2['F8'] = 'г. Муром'
            if ws1['F8'].value == None:
                ws2['F8'] = 'г. Муром'
                if ws1['F8'].value == None:
                    ws2['F8'] = 'г. Муром'
                else:
                    ws2.cell(row=8, column=6).value = ws1['F8'].value
            else:
                if len(ws1['F8'].value) == 0:
                    ws2['F8'] = 'г. Муром'
                else:
                    ws2.cell(row=8, column=6).value = ws1['F8'].value

        elif city_word == 'Ковров':
            ws2['F8'] = 'г. Ковров'
            if ws1['F8'].value == None:
                ws2['F8'] = 'г. Ковров'
                if ws1['F8'].value == None:
                    ws2['F8'] = 'г. Ковров'
                else:
                    ws2.cell(row=8, column=6).value = ws1['F8'].value
            else:
                if len(ws1['F8'].value) == 0:
                    ws2['F8'] = 'г. Ковров'
                else:
                    ws2.cell(row=8, column=6).value = ws1['F8'].value


        ws2['B10'] = 'Заказчик:'
        ws2.merge_cells(start_row=10, start_column=2, end_row=10, end_column=4)
        ws2['B10'].font = fonts.font4

        # Имя заказчика
        ws2.cell(row=10, column=6).value = ws1['F11'].value
        ws2.merge_cells('F10:L10')
        ws2['F10'].font = fonts.font1
        borders.set_border(ws2, 'F10:L10')

        ws2['N9'] = 'Телефоны:'
        ws2.merge_cells(start_row=9, start_column=14, end_row=10, end_column=15)
        ws2['N9'].font = fonts.font2

        # Телефон заказчика
        ws2.cell(row=9, column=17).value = ws1['Q10'].value
        ws2.merge_cells(start_row=9, start_column=17, end_row=10, end_column=27)
        ws2['Q9'].font = fonts.font5
        borders.set_border(ws2, 'Q10:AA10')

        ws2['B12'] = 'Доставка производится:     дата:'
        ws2.merge_cells('B12:I12')
        ws2['B12'].font = fonts.font2

        # Дата доставки
        ws2.cell(row=12, column=11).value = ws1['L13'].value
        ws2.merge_cells('K12:Q12')
        ws2['K12'].font = fonts.font1
        borders.set_border(ws2, 'K12:Q12')

        ws2['S12'] = 'время:'
        ws2.merge_cells('S12:U12')
        ws2['S12'].font = fonts.font2

        ws2['W12'] = 'с 12 до 19'
        ws2.merge_cells('W12:AB12')
        ws2['W12'].font = fonts.font1
        borders.set_border(ws2, 'W12:AB12')

        ws2['B14'] = 'Примечания:'
        ws2.merge_cells('B14:E14')
        ws2['B14'].font = fonts.font2

        # Условия доставки
        ws2['F14'] = 'Самовывоз'
        ws2.merge_cells('F14:Z14')
        ws2['F14'].font = fonts.font1
        borders.set_border(ws2, 'F14:Z14')
        if ws1['F15'].value == None:
            if ws1['F15'].value == ws2['F14'].value:
                ws2['F14'] = 'Самовывоз'
            elif ws1['F15'].value == 'null':
                ws2['F14'] = 'Самовывоз'
            else:
                ws2.cell(row=14, column=6).value = ws1['F15'].value
        elif ws1['F15'].value == 'null':
            ws2['F14'] = 'Самовывоз'
        else:
            if len(ws1['F15'].value) == 0:
                ws2['F14'] = 'Самовывоз'
            elif ws1['F15'].value == 'null':
                ws2['F14'] = 'Самовывоз'
            else:
                ws2.cell(row=14, column=6).value = ws1['F15'].value

        # Обьединение ячеек в таблице заказа
        for k in range(15, 31):
            ws2.merge_cells(start_row=k + 1, start_column=3, end_row=k + 1, end_column=13)
            ws2.merge_cells(start_row=k + 1, start_column=15, end_row=k + 1, end_column=17)
            ws2.merge_cells(start_row=k + 1, start_column=18, end_row=k + 1, end_column=21)
            ws2.merge_cells(start_row=k + 1, start_column=22, end_row=k + 1, end_column=25)

        # Границы в таблице заказа
        for col_cells in ws2.iter_cols(min_col=27, max_col=27, min_row=16, max_row=31):
            for cell in col_cells:
                cell.border = Border(left=Side(border_style='medium', color='000000'))

        for col_cells in ws2.iter_cols(min_col=1, max_col=1, min_row=16, max_row=31):
            for cell in col_cells:
                cell.border = Border(right=Side(border_style='medium', color='000000'))

        for col_cells in ws2.iter_cols(min_col=2, max_col=26, min_row=15, max_row=15):
            for cell in col_cells:
                cell.border = Border(bottom=Side(border_style='medium', color='000000'))

        for col_cells in ws2.iter_cols(min_col=2, max_col=26, min_row=16, max_row=29):
            for cell in col_cells:
                cell.border = Border(bottom=Side(border_style='thin', color='000000'),
                                     right=Side(border_style='thin', color='000000'))

        for col_cells in ws2.iter_cols(min_col=2, max_col=26, min_row=30, max_row=30):
            for cell in col_cells:
                cell.border = Border(bottom=Side(border_style='medium', color='000000'),
                                     right=Side(border_style='thin', color='000000'))

        for col_cells in ws2.iter_cols(min_col=2, max_col=26, min_row=31, max_row=31):
            for cell in col_cells:
                cell.border = Border(bottom=Side(border_style='medium', color='000000'),
                                     right=Side(border_style='thin', color='000000'))

        # Выравнивание ячеек в таблице заказа
        for row in ws2.iter_cols(min_col=14, max_col=26, min_row=16, max_row=30):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        # пп 1 - 14
        for row in ws2.iter_cols(min_col=2, max_col=2, min_row=16, max_row=30):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws2.iter_cols(min_col=3, max_col=13, min_row=17, max_row=30):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for row in ws2.iter_cols(min_col=14, max_col=26, min_row=31, max_row=31):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Шрифт в таблице заказа
        for row in ws2.iter_cols(min_col=3, max_col=26, min_row=17, max_row=30):
            for cell in row:
                cell.font = fonts.font3

        for row in ws2.iter_cols(min_col=3, max_col=26, min_row=31, max_row=31):
            for cell in row:
                cell.font = fonts.font6

        # пп 1 - 14
        for row in ws2.iter_cols(min_col=2, max_col=2, min_row=16, max_row=30):
            for cell in row:
                cell.font = fonts.font6

        # наименование - примечание
        for row in ws2.iter_cols(min_col=3, max_col=26, min_row=16, max_row=16):
            for cell in row:
                cell.font = fonts.font6

        ws2['B16'] = 'пп'
        ws2['B17'] = int('1')
        ws2['B18'] = int('2')
        ws2['B19'] = int('3')
        ws2['B20'] = int('4')
        ws2['B21'] = int('5')
        ws2['B22'] = int('6')
        ws2['B23'] = int('7')
        ws2['B24'] = int('8')
        ws2['B25'] = int('9')
        ws2['B26'] = int('10')
        ws2['B27'] = int('11')
        ws2['B28'] = int('12')
        ws2['B29'] = int('13')
        ws2['B30'] = int('14')

        ws2['C16'] = 'наименование'
        currentCell = ws2.cell(row=16, column=3)
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws2['N16'] = 'к-во'
        ws2['O16'] = 'мест'
        ws2['R16'] = 'цена'
        ws2['V16'] = 'сумма'
        ws2['Z16'] = 'примечание'
        ws2['C31'] = 'итого:'
        currentCell = ws2.cell(row=31, column=3)
        currentCell.alignment = Alignment(horizontal='right', vertical='center')

        # Первая позиция
        def pp1():
            for i in ws2:
                if ws1['B18'].value == 1:
                    # Наименование
                    ws2.cell(row=17, column=3).value = ws1['C18'].value
                    # К-во
                    ws2.cell(row=17, column=14).value = ws1['N18'].value
                    # мест
                    ws2.cell(row=17, column=15).value = ws1['O18'].value
                    # цена
                    ws2.cell(row=17, column=18).value = ws1['R18'].value
                    # сумма
                    ws2.cell(row=17, column=22).value = ws1['V18'].value
                    # примечание
                    ws2.cell(row=17, column=26).value = ws1['Z18'].value

                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C18'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R18'] = c
                            ws2['V18'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R18'] = c
                            ws2['V18'] = c
                    else:
                        break

        pp1()

        # Вторая позиция
        def pp2():
            for n in ws2:
                if ws1['B19'].value == 2:
                    # Наименование
                    ws2.cell(row=18, column=3).value = ws1['C19'].value
                    # К-во
                    ws2.cell(row=18, column=14).value = ws1['N19'].value
                    # мест
                    ws2.cell(row=18, column=15).value = ws1['O19'].value
                    # цена
                    ws2.cell(row=18, column=18).value = ws1['R19'].value
                    # сумма
                    ws2.cell(row=18, column=22).value = ws1['V19'].value
                    # примечание
                    ws2.cell(row=18, column=26).value = ws1['Z19'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C18'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R18'] = c
                            ws2['V18'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R18'] = c
                            ws2['V18'] = c
                    else:
                        break

        if ws1['C18'].value == None:
            ws2['C18'] = ''
        elif ws2['C17'].value == ws1['C18'].value:
            pp2()

        # Третья позиция
        def pp3():
            for n in ws2:
                if ws1['B20'].value == 3:
                    # Наименование
                    ws2.cell(row=19, column=3).value = ws1['C20'].value
                    # К-во
                    ws2.cell(row=19, column=14).value = ws1['N20'].value
                    # мест
                    ws2.cell(row=19, column=15).value = ws1['O20'].value
                    # цена
                    ws2.cell(row=19, column=18).value = ws1['R20'].value
                    # сумма
                    ws2.cell(row=19, column=22).value = ws1['V20'].value
                    # примечание
                    ws2.cell(row=19, column=26).value = ws1['Z20'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C19'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R19'] = c
                            ws2['V19'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R19'] = c
                            ws2['V19'] = c
                    else:
                        break

        if ws1['C19'].value == None:
            ws2['C19'] = ''
        elif ws2['C18'].value == ws1['C19'].value:
            pp3()

        # Четвертая позиция
        def pp4():
            for n in ws2:
                if ws1['B21'].value == 4:
                    # Наименование
                    ws2.cell(row=20, column=3).value = ws1['C21'].value
                    # К-воow
                    ws2.cell(row=20, column=14).value = ws1['N21'].value
                    # мест
                    ws2.cell(row=20, column=15).value = ws1['O21'].value
                    # цена
                    ws2.cell(row=20, column=18).value = ws1['R21'].value
                    # сумма
                    ws2.cell(row=20, column=22).value = ws1['V21'].value
                    # примечание
                    ws2.cell(row=20, column=26).value = ws1['Z21'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C20'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R20'] = c
                            ws2['V20'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R20'] = c
                            ws2['V20'] = c
                    else:
                        break

        if ws1['C20'].value == None:
            ws2['C20'] = ''
        elif ws2['C19'].value == ws1['C20'].value:
            pp4()

        # Пятая позиция
        def pp5():
            for n in ws2:
                if ws1['B22'].value == 5:
                    # Наименование
                    ws2.cell(row=21, column=3).value = ws1['C22'].value
                    # К-во
                    ws2.cell(row=21, column=14).value = ws1['N22'].value
                    # мест
                    ws2.cell(row=21, column=15).value = ws1['O22'].value
                    # цена
                    ws2.cell(row=21, column=18).value = ws1['R22'].value
                    # сумма
                    ws2.cell(row=21, column=22).value = ws1['V22'].value
                    # примечание
                    ws2.cell(row=21, column=26).value = ws1['Z22'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C21'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R21'] = c
                            ws2['V21'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R21'] = c
                            ws2['V21'] = c
                    else:
                        break

        if ws1['C21'].value == None:
            ws2['C21'] = ''
        elif ws2['C20'].value == ws1['C21'].value:
            pp5()

        # Шестая позиция
        def pp6():
            for n in ws2:
                if ws1['B23'].value == 6:
                    # Наименование
                    ws2.cell(row=22, column=3).value = ws1['C23'].value
                    # К-во
                    ws2.cell(row=22, column=14).value = ws1['N23'].value
                    # мест
                    ws2.cell(row=22, column=15).value = ws1['O23'].value
                    # цена
                    ws2.cell(row=22, column=18).value = ws1['R23'].value
                    # сумма
                    ws2.cell(row=22, column=22).value = ws1['V23'].value
                    # примечание
                    ws2.cell(row=22, column=26).value = ws1['Z23'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C22'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R22'] = c
                            ws2['V22'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R22'] = c
                            ws2['V22'] = c
                    else:
                        break

        if ws1['C22'].value == None:
            ws2['C22'] = ''
        elif ws2['C21'].value == ws1['C22'].value:
            pp6()

        # Седьмая позиция
        def pp7():
            for n in ws2:
                if ws1['B24'].value == 7:
                    # Наименование
                    ws2.cell(row=23, column=3).value = ws1['C24'].value
                    # К-во
                    ws2.cell(row=23, column=14).value = ws1['N24'].value
                    # мест
                    ws2.cell(row=23, column=15).value = ws1['O24'].value
                    # цена
                    ws2.cell(row=23, column=18).value = ws1['R24'].value
                    # сумма
                    ws2.cell(row=23, column=22).value = ws1['V24'].value
                    # примечание
                    ws2.cell(row=23, column=26).value = ws1['Z24'].value
                else:

                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C23'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R23'] = c
                            ws2['V23'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R23'] = c
                            ws2['V23'] = c
                    else:
                        break

        if ws1['C23'].value == None:
            ws2['C23'] = ''
        elif ws2['C22'].value == ws1['C23'].value:
            pp7()

        # Восьмая позиция
        def pp8():
            for n in ws2:
                if ws1['B25'].value == 8:
                    # Наименование
                    ws2.cell(row=24, column=3).value = ws1['C25'].value
                    # К-во
                    ws2.cell(row=24, column=14).value = ws1['N25'].value
                    # мест
                    ws2.cell(row=24, column=15).value = ws1['O25'].value
                    # цена
                    ws2.cell(row=24, column=18).value = ws1['R25'].value
                    # сумма
                    ws2.cell(row=24, column=22).value = ws1['V25'].value
                    # примечание
                    ws2.cell(row=24, column=26).value = ws1['Z25'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C24'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R24'] = c
                            ws2['V24'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R24'] = c
                            ws2['V24'] = c
                    else:
                        break

        if ws1['C24'].value == None:
            ws2['C24'] = ''
        elif ws2['C23'].value == ws1['C24'].value:
            pp8()

        # Девятая позиция
        def pp9():
            for n in ws2:
                if ws1['B26'].value == 9:
                    # Наименование
                    ws2.cell(row=25, column=3).value = ws1['C26'].value
                    # К-во
                    ws2.cell(row=25, column=14).value = ws1['N26'].value
                    # мест
                    ws2.cell(row=25, column=15).value = ws1['O26'].value
                    # цена
                    ws2.cell(row=25, column=18).value = ws1['R26'].value
                    # сумма
                    ws2.cell(row=25, column=22).value = ws1['V26'].value
                    # примечание
                    ws2.cell(row=25, column=26).value = ws1['Z26'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C25'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R25'] = c
                            ws2['V25'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R25'] = c
                            ws2['V25'] = c
                    else:
                        break

        if ws1['C25'].value == None:
            ws2['C25'] = ''
        elif ws2['C24'].value == ws1['C25'].value:
            pp9()

        # Десятая позиция
        def pp10():
            for n in ws2:
                if ws1['B27'].value == 10:
                    # Наименование
                    ws2.cell(row=26, column=3).value = ws1['C27'].value
                    # К-во
                    ws2.cell(row=26, column=14).value = ws1['N27'].value
                    # мест
                    ws2.cell(row=26, column=15).value = ws1['O27'].value
                    # цена
                    ws2.cell(row=26, column=18).value = ws1['R27'].value
                    # сумма
                    ws2.cell(row=26, column=22).value = ws1['V27'].value
                    # примечание
                    ws2.cell(row=26, column=26).value = ws1['Z27'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C26'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R26'] = c
                            ws2['V26'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R26'] = c
                            ws2['V26'] = c
                    else:
                        break

        if ws1['C26'].value == None:
            ws2['C26'] = ''
        elif ws2['C25'].value == ws1['C26'].value:
            pp10()

        # Одиннадцатая позиция
        def pp11():
            for n in ws2:
                if ws1['B28'].value == 11:
                    # Наименование
                    ws2.cell(row=27, column=3).value = ws1['C28'].value
                    # К-во
                    ws2.cell(row=27, column=14).value = ws1['N28'].value
                    # мест
                    ws2.cell(row=27, column=15).value = ws1['O28'].value
                    # цена
                    ws2.cell(row=27, column=18).value = ws1['R28'].value
                    # сумма
                    ws2.cell(row=27, column=22).value = ws1['V28'].value
                    # примечание
                    ws2.cell(row=27, column=26).value = ws1['Z28'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C27'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R27'] = c
                            ws2['V27'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R27'] = c
                            ws2['V27'] = c
                    else:
                        break

        if ws1['C27'].value == None:
            ws2['C27'] = ''
        elif ws2['C26'].value == ws1['C27'].value:
            pp11()

        # Двеннадцатая позиция
        def pp12():
            for n in ws2:
                if ws1['B29'].value == 12:
                    # Наименование
                    ws2.cell(row=28, column=3).value = ws1['C29'].value
                    # К-во
                    ws2.cell(row=28, column=14).value = ws1['N29'].value
                    # мест
                    ws2.cell(row=28, column=15).value = ws1['O29'].value
                    # цена
                    ws2.cell(row=28, column=18).value = ws1['R29'].value
                    # сумма
                    ws2.cell(row=28, column=22).value = ws1['V29'].value
                    # примечание
                    ws2.cell(row=28, column=26).value = ws1['Z29'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C28'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R28'] = c
                            ws2['V28'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R28'] = c
                            ws2['V28'] = c
                    else:
                        break

        if ws1['C28'].value == None:
            ws2['C28'] = ''
        elif ws2['C27'].value == ws1['C28'].value:
            pp12()

        # Тринадцатая позиция
        def pp13():
            for n in ws2:
                if ws1['B30'].value == 13:
                    # Наименование
                    ws2.cell(row=29, column=3).value = ws1['C30'].value
                    # К-во
                    ws2.cell(row=29, column=14).value = ws1['N30'].value
                    # мест
                    ws2.cell(row=29, column=15).value = ws1['O30'].value
                    # цена
                    ws2.cell(row=29, column=18).value = ws1['R30'].value
                    # сумма
                    ws2.cell(row=29, column=22).value = ws1['V30'].value
                    # примечание
                    ws2.cell(row=29, column=26).value = ws1['Z30'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C29'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R29'] = c
                            ws2['V29'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R29'] = c
                            ws2['V29'] = c
                    else:
                        break

        if ws1['C29'].value == None:
            ws2['C29'] = ''
        elif ws2['C28'].value == ws1['C29'].value:
            pp13()

        # Четырнадцатая позиция
        def pp14():
            for n in ws2:
                if ws1['B31'].value == 14:
                    # Наименование
                    ws2.cell(row=30, column=3).value = ws1['C31'].value
                    # К-во
                    ws2.cell(row=30, column=14).value = ws1['N31'].value
                    # мест
                    ws2.cell(row=30, column=15).value = ws1['O31'].value
                    # цена
                    ws2.cell(row=30, column=18).value = ws1['R31'].value
                    # сумма
                    ws2.cell(row=30, column=22).value = ws1['V31'].value
                    # примечание
                    ws2.cell(row=30, column=26).value = ws1['Z31'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C30'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R30'] = c
                            ws2['V30'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R30'] = c
                            ws2['V30'] = c
                    else:
                        break

        if ws1['C30'].value == None:
            ws2['C30'] = ''
        elif ws2['C29'].value == ws1['C30'].value:
            pp14()

        ws2['N31'] = '=SUM(N17:N30)'
        ws2['O31'] = '=SUM(O17:O30)'
        ws2['V31'] = '=SUM(V17:V30)'

        ws2['B33'] = 'АКТ ПРИЕМА-ПЕРЕДАЧИ '
        ws2.merge_cells('B33:Z33')
        currentCell = ws2.cell(row=33, column=2)
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        ws2['B33'].font = fonts.font6

        ws2['B34'] = 'Продавец сдал, а Покупатель принял Товар в ассортименте и количестве, ' \
                     'согласованном в заказе Покупателя. При достаточном освещении и пространстве ' \
                     'для осмотра Покупатель непосредственно ознакомлен с техническими характеристиками, ' \
                     'свойствами, функциями и габаритами, которые полностью соответствуют описанию Товара, ' \
                     'предоставленному до оформления заказа. Продавцом передан Товар надлежащего качества, ' \
                     'с полным комплектом принадлежностей и сопроводительных документов (гарантийный талон, ' \
                     'инструкция на русском языке и другие документы к Товару, предоставление которых ' \
                     'предусмотрено для Товара данного вида), претензий к упаковке, комплектности, ' \
                     'внешнему виду Покупатель не имеет, механических повреждений нет. Установленный ' \
                     'изготовителем комплект принадлежностей, технический паспорт или иной заменяющий ' \
                     'его документ, инструкция по эксплуатации на русском языке и другие относящиеся ' \
                     'к нему документы Покупателем получены полностью.  Необходимая и достоверная ' \
                     'информация об изготовителе (Продавце) доведена до сведения Покупателя полностью. ' \
                     'Гарантийный талон получен и оформлен надлежащим образом. Продавец полностью исполнил ' \
                     'свою обязанность по передаче Товара Покупателю.'
        ws2.merge_cells(start_row=34, start_column=2, end_row=43, end_column=26)
        currentCell = ws2.cell(row=34, column=2)
        currentCell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws2['B34'].font = fonts.font7

        for row in ws2.iter_cols(min_col=2, max_col=26, min_row=44, max_row=48):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)

        ws2['B44'] = '8. ПОДПИСИ СТОРОН'
        ws2.merge_cells('B44:Z44')
        ws2['B44'].font = fonts.font3

        ws2['B45'] = 'Настоящий Договор составлен в двух экземплярах, первый остается ' \
                     'у представителя Продавца, второй экземпляр выдается Покупателю. ' \
                     'Договор вступает в силу с момента его подписания сторонами и действует ' \
                     'до момента надлежащего исполнения сторонами принятых на себя обязательств.'
        ws2.merge_cells(start_row=45, start_column=2, end_row=47, end_column=26)
        ws2['B45'].font = fonts.font7

        ws2['B48'] = 'С условиями настоящего договора ознакомлен до момента его подписания и полностью согласен'
        ws2.merge_cells('B48:Z48')
        ws2['B48'].font = fonts.font7

        ws2['K49'] = 'Представитель Продавца'
        ws2.merge_cells('K49:M49')
        currentCell = ws2.cell(row=49, column=11)
        currentCell.alignment = Alignment(horizontal='right', vertical='bottom')
        ws2['K49'].font = fonts.font3

        borders.set_border(ws2, 'N49:U49')

        for row in ws2.iter_cols(min_col=14, max_col=26, min_row=49, max_row=50):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='bottom')

        # ФИО после подписи
        ws2.merge_cells('X49:Z49')
        ws2['X49'].font = fonts.font3
        borders.set_border(ws2, 'X49:Z49')

        ws2['N50'] = '(Подпись)'
        ws2.merge_cells('N50:V50')
        ws2['N50'].font = fonts.font7

        ws2['X50'] = '(ф.и.о.)'
        ws2.merge_cells('X50:Z50')
        ws2['X50'].font = fonts.font7

        ws2['B52'] = 'Покупатель'
        ws2.merge_cells('B52:D52')
        currentCell = ws2.cell(row=52, column=2)
        currentCell.alignment = Alignment(horizontal='left', vertical='bottom')
        ws2['B52'].font = fonts.font1

        borders.set_border(ws2, 'E52:H52')
        borders.set_border(ws2, 'K52:Q52')

        ws2['R52'] = 'Дата '
        ws2.merge_cells('R52:T52')
        currentCell = ws2.cell(row=52, column=18)
        currentCell.alignment = Alignment(horizontal='right', vertical='bottom')
        ws2['R52'].font = fonts.font1

        ws2['U52'] = '«       »                                             2022г.'
        ws2.merge_cells('U52:Z52')
        currentCell = ws2.cell(row=52, column=21)
        currentCell.alignment = Alignment(horizontal='left', vertical='bottom')
        ws2['U52'].font = fonts.font1
        borders.set_border(ws2, 'U52:Z52')

        ws2['E53'] = ' (Подпись)'
        ws2.merge_cells('E53:H53')
        currentCell = ws2.cell(row=53, column=5)
        currentCell.alignment = Alignment(horizontal='center', vertical='bottom')
        ws2['E53'].font = fonts.font6

        ws2['K53'] = ' (ф.и.о.)'
        ws2.merge_cells('K53:Q53')
        currentCell = ws2.cell(row=53, column=11)
        currentCell.alignment = Alignment(horizontal='center', vertical='bottom')
        ws2['K53'].font = fonts.font6

        borders.set_border5(ws2, 'A54:Z54')

        img = Image('logo.jpg')
        ws2.add_image(img, 'B2')

        # Вторая страница
        ws2.column_dimensions['AC'].width = 0.05
        ws2.column_dimensions['AD'].width = 5.71 + 0.71
        ws2.column_dimensions['AE'].width = 3.14 + 0.71
        ws2.column_dimensions['AF'].width = 7 + 0.71
        ws2.column_dimensions['AG'].width = 1.14 + 0.71
        ws2.column_dimensions['AH'].width = 0.75 + 0.71
        ws2.column_dimensions['AI'].width = 3.71 + 0.71
        ws2.column_dimensions['AJ'].width = 3 + 0.71
        ws2.column_dimensions['AK'].width = 0.17 + 0.71
        ws2.column_dimensions['AL'].width = 16.71 + 0.71
        ws2.column_dimensions['AM'].width = 0.42 + 0.71
        ws2.column_dimensions['AN'].width = 5.14 + 0.71
        ws2.column_dimensions['AO'].width = 3.43 + 0.71
        ws2.column_dimensions['AP'].width = 0.42 + 0.71
        ws2.column_dimensions['AQ'].width = 0.42 + 0.71
        ws2.column_dimensions['AR'].width = 0.75 + 0.71
        ws2.column_dimensions['AS'].width = 2.43 + 0.71
        ws2.column_dimensions['AT'].width = 0.17 + 0.71
        ws2.column_dimensions['AU'].width = 4 + 0.71
        ws2.column_dimensions['AV'].width = 2.29 + 0.71
        ws2.column_dimensions['AW'].width = 0.33 + 0.71
        ws2.column_dimensions['AX'].width = 2.43 + 0.71
        ws2.column_dimensions['AY'].width = 3.43 + 0.71
        ws2.column_dimensions['AZ'].width = 0.92 + 0.71
        ws2.column_dimensions['BA'].width = 7 + 0.71
        ws2.column_dimensions['BB'].width = 10 + 0.71
        ws2.column_dimensions['BC'].width = 8.43 + 0.71
        ws2.column_dimensions['BD'].width = 8.43 + 0.71

        ws2.merge_cells(start_row=2, start_column=37, end_row=3, end_column=53)
        currentCell = ws2.cell(row=2, column=37)
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        ws2['AK2'].font = fonts.font1

        for row in ws2.iter_cols(min_col=55, max_col=56, min_row=2, max_row=6):
            for cell in row:
                cell.alignment = Alignment(horizontal='right', vertical='center')

        for row in ws2.iter_cols(min_col=55, max_col=56, min_row=2, max_row=6):
            for cell in row:
                cell.font = fonts.font2

        ws2['BC2'] = 'Отдел продаж'
        ws2.merge_cells(start_row=2, start_column=55, end_row=2, end_column=56)

        ws2['BC3'] = '8 (800) 555-87-78'
        ws2.merge_cells(start_row=3, start_column=55, end_row=3, end_column=56)

        ws2['BC5'] = 'Служба доставки'
        ws2.merge_cells(start_row=5, start_column=55, end_row=5, end_column=56)

        ws2.merge_cells(start_row=6, start_column=55, end_row=6, end_column=56)

        ws2['AD7'] = 'ПИПИПИПИПИПИПИПИ, именуемый в дальнейшем ' \
                     'Продавец, и _____________________________________________________, ' \
                     'именуемый в дальнейшем Покупатель, заключили настоящий  договор'
        if city_word == 'Владимир':
            p = ws2['AD7'].value
            px = p.replace('ПИПИПИПИПИПИПИПИ', 'ИП Трофимова Т.П. - TechPort.Vladimir')
            ws2['AD7'] = px
        elif city_word == 'Иваново':
            p = ws2['AD7'].value
            px = p.replace('ПИПИПИПИПИПИПИПИ', 'ИП Данилов С.Ю. - TechPort.Ivanovo')
            ws2['AD7'] = px
        elif city_word == 'Муром':
            p = ws2['AD7'].value
            px = p.replace('ПИПИПИПИПИПИПИПИ', 'ИП Трофимова Т.П. - TechPort.Murom')
            ws2['AD7'] = px
        elif city_word == 'Ковров':
            p = ws2['AD7'].value
            px = p.replace('ПИПИПИПИПИПИПИПИ', 'ИП Трофимова Т.П. - TechPort.Kovrov')
            ws2['AD7'] = px

        ws2.merge_cells('AD7:BD7')
        currentCell = ws2.cell(row=7, column=30)
        currentCell.alignment = Alignment(horizontal='left', vertical='top')
        ws2['AD7'].font = fonts.font8

        ws2['AD8'] = '1. ПРЕДМЕТ ДОГОВОРА' \
                     '                                                                       ' \
                     '                                                                       ' \
                     '                                                                         ' \
                     '                                                                        ' \
                     '1.1.  По настоящему договору ' \
                     'Продавец обязуется передать в собственность ' \
                     'Покупателя Товар в количестве, ассортименте и в сроки, предусмотренные настоящим ' \
                     'Договором, а Покупатель обязуется оплатить Товар по цене, предусмотренной в Договоре. ' \
                     'Заключая настоящий договор, Покупатель соглашается с тем, что Продавец предложил ему ' \
                     'полную информацию о товаре.' \
                     '                                                                     ' \
                     '                                                                     ' \
                     '                                                                     ' \
                     '                                                                     ' \
                     '1.2. Покупатель обязуется лично осмотреть Товар, после чего в случае отсутствия ' \
                     'выявленных недостатков принять Товар и уплатить за него согласованную в договоре цену.' \
                     '                                                                       ' \
                     '                                                                       ' \
                     '                                                                         ' \
                     '                                                                        ' \
                     '2. ПРАВА И ОБЯЗАННОСТИ СТОРОН' \
                     '                                                                       ' \
                     '                                                                       ' \
                     '                                                                         ' \
                     '                                                                        ' \
                     '2.1. Продавец обязан:' \
                     '                                                                       ' \
                     '                                                                       ' \
                     '                                                                         ' \
                     '                                                                        ' \
                     '2.1.1. Предоставить Покупателю Товар надлежащего качества в количестве и ' \
                     'ассортименте, соответствующем заказу Покупателя, в согласованные в нем сроки ' \
                     'путем доставки по указанному Покупателем адресу.' \
                     '                                                                       ' \
                     '                                                                       ' \
                     '                                                                         ' \
                     '                                                                        ' \
                     '2.1.2. Предоставить Покупателю необходимую и достоверную информацию о:' \
                     '- Продавце и изготовителе Товара; - предприятиях сервисного обслуживания, ' \
                     'уполномоченных на проведение гарантийного ремонта, - порядке, условиях и сроках' \
                     ' возврата Товара Продавцу. Своей подписью в настоящем Договоре Покупатель ' \
                     'подтверждает получение в технической документации, прилагаемой к Товару, на' \
                     ' этикетке, маркировке, упаковке Товара или иным способом, следующих сведений ' \
                     'о Товаре в момент его доставки: сведения об основных потребительских свойствах ' \
                     'Товара (технические характеристики, функции, габариты, комплектация, условия ' \
                     'подключения, наладки, пуска в эксплуатацию, условия правильной и безопасной ' \
                     'эксплуатации, правила и условия гарантийного обслуживания, информация об ' \
                     'изготовителе/производителе Товара и т. д.); сведения о сроке службы Товара.' \
                     '                                                                       ' \
                     '                                                                       ' \
                     '                                                                         ' \
                     '                                                                        ' \
                     '2.2. Покупатель обязан:                                                ' \
                     '                                                                       ' \
                     '                                                                       ' \
                     '                                                                         ' \
                     '                                                                        ' \
                     '2.2.1. Находиться по адресу доставки лично весь интервал времени, согласованный ' \
                     'настоящим договором, для непосредственного осмотра Товара, его принятия и оплаты ' \
                     'в присутствии уполномоченного представителя Продавца.                     ' \
                     '                                                                          ' \
                     '                                                                          ' \
                     '                                                                            ' \
                     '                                                                           ' \
                     '2.2.2. Обеспечить условия и место, необходимые и достаточные для осмотра ' \
                     'Товара до подписания Акта приема-передачи.' \
                     '                                                                       ' \
                     '                                                                       ' \
                     '                                                                         ' \
                     '                                                                        ' \
                     '2.2.3. После осмотра Товара в случае отсутствия недостатков подписать Акт ' \
                     'приема-передачи и уплатить согласованную сторонами цену за Товар ' \
                     '(за исключением случая его полной предварительной оплаты).' \
                     '                                                                       ' \
                     '                                                                       ' \
                     '                                                                         ' \
                     '                                                                        ' \
                     '2.2.4. Осуществить ' \
                     'сборку и/или установку (подключение) технически сложного Товара с привлечением ' \
                     'уполномоченной изготовителем (Продавцом) организации, самостоятельная сборка и/или ' \
                     'подключение которого Покупателем в соответствии со стандартами или технической ' \
                     'документацией, прилагаемой к Товару (технический паспорт, инструкция по эксплуатации, ' \
                     'гарантийный талон), не допускается. В случаях, если Покупатель не обращался за ' \
                     'подключением и пуском в эксплуатацию Товара, указанного в настоящем пункте, в ' \
                     'уполномоченные на их совершение организации и они были произведены лицами, не ' \
                     'имеющими необходимой квалификации, Покупатель утрачивает право предъявления Продавцу' \
                     ' претензий о ненадлежащей работе и/или неисправности Товара, а также право на его гарантийный ремонт.' \
                     '                                                                       ' \
                     '                                                                       ' \
                     '                                                                         ' \
                     '                                                                        ' \
                     '2.3. Покупатель вправе отказаться от принятия и оплаты заказанного Товара, если ' \
                     'характеристики Товара имеют расхождения с предоставленным Продавцом описанием ' \
                     'Товара или в случае, если при осмотре Покупателем Товара выявлены недостатки ' \
                     '(механические повреждения, недостатки качества), не согласованные с Покупателем' \
                     ' при оформлении заказа.                                     ' \
                     '                                                                          ' \
                     '                                                                          ' \
                     '                                                                            ' \
                     '                                                                           ' \
                     '2.4. Товар является качественным, если он пригоден для использования в ' \
                     'целях, для которых Товар данного рода обычно используется.               ' \
                     '                                                                       ' \
                     '                                                                       ' \
                     '                                                                         ' \
                     '                                                                        ' \
                     '3. ПЕРЕХОД ПРАВА СОБСТВЕННОСТИ НА ТОВАР, РИСКА ЕГО СЛУЧАЙНОЙ ГИБЕЛИ ИЛИ ПОВРЕЖДЕНИЯ' \
                     'Право собственности на Товар переходит к Покупателю с момента подписания Акта' \
                     ' приема-передачи. После перехода права собственности на Товар Продавец считается' \
                     ' исполнившим свою обязанность по передаче Товара Покупателю. Датой передачи Товара ' \
                     'Покупателю считается дата подписания Акта приема-передачи Товара. Риск случайной ' \
                     'гибели или случайного повреждения Товара переходит на Покупателя после перехода права ' \
                     'собственности на Товар. С указанного момента Покупатель не имеет права предъявлять ' \
                     'претензии по внешнему виду, комплектации и механическим повреждениям Товара.' \
                     '                                                                       ' \
                     '                                                                       ' \
                     '                                                                         ' \
                     '                                                                        ' \
                     '4. ОБМЕН (ВОЗВРАТ) ТОВАРА ' \
                     '                                                                       ' \
                     '                                                                       ' \
                     '                                                                         ' \
                     '                                                                        ' \
                     '4.1. Обмен или возврат Товара надлежащего качества осуществляется в следующем' \
                     ' порядке: Покупатель вправе отказаться от Товара в любое время до его передачи, ' \
                     'а после передачи — в течение 7 (Семи) дней. Возврат Товара надлежащего качества ' \
                     'возможен в случае, если сохранены его товарный вид, потребительские свойства, ' \
                     'документ, подтверждающий факт и условия покупки. В случае, если Покупателем не ' \
                     'сохранены товарный вид и/или потребительские свойства Товара (в том числе в случае' \
                     ' появления на момент передачи Товара от Покупателя к Продавцу не оговоренных при ' \
                     'заключении настоящего договора дефектов: внешний вид, неполная комплектация и т. д. ' \
                     'и иных недостатков, отсутствовавших на момент заключения настоящего Договора), ' \
                     'Продавец вправе отказаться от приема Товара от Покупателя. В случае утраты и/или ' \
                     'порчи Покупателем упаковки, не разрушаемой при ее вскрытии и обычной для упаковки ' \
                     'Товаров такого рода при их продаже, Покупатель лишается права на возврат Товара по ' \
                     'причине утраты товарного вида, даже если сам Товар не поврежден. Товар, поставляемый' \
                     ' в разрушаемой упаковке (блистерная и другие виды упаковки), после ее вскрытия обмену' \
                     ' и возврату не подлежит по причине утраты товарного вида, даже если сам Товар не поврежден.' \
                     '                                                                             ' \
                     '                                                                             ' \
                     '                                                                             ' \
                     '                                                                             ' \
                     '4.2.Возврат Товара надлежащего качества осуществляется' \
                     f' по адресу: АРАРАРАРАРАРАР,  при соблюдении Покупателем условий, ' \
                     'изложенных в настоящем Договоре. При соблюдении Покупателем условий возврата Товара,' \
                     ' Продавец в сроки, установленные законодательством РФ возвращает Покупателю сумму, ' \
                     'уплаченную Покупателем за Товар. Возврат данной суммы осуществляется в наличном или' \
                     ' безналичном порядке, аналогичном порядку ее уплаты Покупателем.' \
                     '                                                                       ' \
                     '                                                                       ' \
                     '                                                                         ' \
                     '                                                                        ' \
                     '4.3. Покупатель не вправе вернуть Товар в случае, если такой Товар в соответствии' \
                     ' с законодательством РФ не подлежит обмену или возврату.' \
                     '                                                                       ' \
                     '                                                                       ' \
                     '                                                                         ' \
                     '                                                                        ' \
                     '4.4. Доставка возвращенного Товара надлежащего качества производится силами' \
                     ' Продавца за счет Покупателя.' \
                     '                                                                       ' \
                     '                                                                       ' \
                     '                                                                         ' \
                     '                                                                        ' \
                     '4.5. Обмен или возврат Товара ненадлежащего качества осуществляется в' \
                     ' соответствии с законодательством РФ.' \
                     '                                                                       ' \
                     '                                                                       ' \
                     '                                                                         ' \
                     '                                                                        ' \
                     '4.6. Продавец не несет ответственность за перемещение техники, осуществлённое' \
                     ' по частной договоренности с экипажем доставки, если такое перемещение' \
                     ' осуществлено после подписания Покупателем Акта приема-передачи.   ' \
                     '                                                                            ' \
                     '                                                                            ' \
                     '                                                                             ' \
                     '                                                                             ' \
                     '5. Информация об условиях и стоимости услуг по доставке опубликована ' \
                     'на сайте www.techport.ru.                         ' \
                     '                                                                       ' \
                     '                                                                       ' \
                     '                                                                         ' \
                     '                                                                        ' \
                     '6. В случае неисправности транспорта, Продавец оставляет за собой право ' \
                     'изменить дату и время доставки, о чем заблаговременно оповещает Покупателя ' \
                     'и согласует с ним новый срок доставки.'

        if city_word == 'Владимир':
            p = ws2['AD8'].value
            px = p.replace('АРАРАРАРАРАРАР', 'г. Владимир, ул. Ново-Ямская, д.75')
            ws2['AD8'] = px
        elif city_word == 'Иваново':
            p = ws2['AD8'].value
            px = p.replace('АРАРАРАРАРАРАР', 'г. Иваново, ул. Лежневская, д.183')
            ws2['AD8'] = px
        elif city_word == 'Муром':
            p = ws2['AD8'].value
            px = p.replace('АРАРАРАРАРАРАР', 'г. Муром, ул. Московская, д.5')
            ws2['AD8'] = px
        elif city_word == 'Ковров':
            p = ws2['AD8'].value
            px = p.replace('АРАРАРАРАРАРАР', 'г. Владимир, ул. Ново-Ямская, д.75')
            ws2['AD8'] = px

        ws2.merge_cells(start_row=8, start_column=30, end_row=53, end_column=56)
        currentCell = ws2.cell(row=8, column=30)
        currentCell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws2['AD8'].font = fonts.font8

        img2 = Image('logo.jpg')
        ws2.add_image(img2, 'AD2')

        borders.set_border5(ws2, 'AD54:BD54')

        # Вторая половина страницы
        ws2.row_dimensions[55].height = 15
        ws2.row_dimensions[56].height = 9.75
        ws2.row_dimensions[57].height = 10.5
        ws2.row_dimensions[58].height = 9.75
        ws2.row_dimensions[59].height = 9
        ws2.row_dimensions[60].height = 10.5
        ws2.row_dimensions[61].height = 10.5
        ws2.row_dimensions[62].height = 13.5
        ws2.row_dimensions[63].height = 3
        ws2.row_dimensions[64].height = 14
        ws2.row_dimensions[65].height = 3
        ws2.row_dimensions[66].height = 12
        ws2.row_dimensions[67].height = 2.25
        ws2.row_dimensions[68].height = 12
        ws2.row_dimensions[69].height = 5.25
        ws2.row_dimensions[70].height = 11.5
        ws2.row_dimensions[71].height = 11.5
        ws2.row_dimensions[72].height = 11.5
        ws2.row_dimensions[73].height = 11.5
        ws2.row_dimensions[74].height = 11.5
        ws2.row_dimensions[75].height = 11.5
        ws2.row_dimensions[76].height = 11.5
        ws2.row_dimensions[77].height = 11.5
        ws2.row_dimensions[78].height = 11.5
        ws2.row_dimensions[79].height = 11.5
        ws2.row_dimensions[80].height = 11.5
        ws2.row_dimensions[81].height = 11.5
        ws2.row_dimensions[82].height = 11.5
        ws2.row_dimensions[83].height = 11.5
        ws2.row_dimensions[84].height = 11.5
        ws2.row_dimensions[85].height = 11.5
        ws2.row_dimensions[86].height = 2.25
        ws2.row_dimensions[87].height = 10.5
        ws2.row_dimensions[88].height = 8
        ws2.row_dimensions[89].height = 8
        ws2.row_dimensions[90].height = 8
        ws2.row_dimensions[91].height = 8
        ws2.row_dimensions[92].height = 8
        ws2.row_dimensions[93].height = 8
        ws2.row_dimensions[94].height = 8
        ws2.row_dimensions[95].height = 8
        ws2.row_dimensions[96].height = 8
        ws2.row_dimensions[97].height = 6
        ws2.row_dimensions[98].height = 11.25
        ws2.row_dimensions[99].height = 3.75
        ws2.row_dimensions[100].height = 5.25
        ws2.row_dimensions[101].height = 11.25
        ws2.row_dimensions[102].height = 9.75
        ws2.row_dimensions[103].height = 12.75
        ws2.row_dimensions[104].height = 10.5
        ws2.row_dimensions[105].height = 4
        ws2.row_dimensions[106].height = 20
        ws2.row_dimensions[107].height = 9.75

        # Большой текст
        ws2.cell(row=62, column=30).value = ws2['AD8'].value
        ws2.merge_cells(start_row=62, start_column=30, end_row=107, end_column=56)
        currentCell = ws2.cell(row=62, column=30)
        currentCell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws2['AD62'].font = fonts.font8

        for row in ws2.iter_cols(min_col=55, max_col=56, min_row=56, max_row=60):
            for cell in row:
                cell.alignment = Alignment(horizontal='right', vertical='center')

        for row in ws2.iter_cols(min_col=55, max_col=56, min_row=56, max_row=60):
            for cell in row:
                cell.font = fonts.font2

        ws2.merge_cells(start_row=56, start_column=37, end_row=57, end_column=53)
        currentCell = ws2.cell(row=56, column=37)
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        ws2['AK56'].font = fonts.font1

        ws2['X103'] = 'Трофимова Татьяна Павловна'
        if city_word == 'Владимир':
            shapka = 'ИП Трофимова Т.П. - TechPort.Vladimir'
            ws2['AK56'] = f'  {shapka}'
            ws2['BC60'] = '8 (4922) 46-20-57'
        elif city_word == 'Иваново':
            shapka = 'ИП Данилов С.Ю. - TechPort.Ivanovo'
            ws2['AK56'] = f'  {shapka}'
            ws2['BC60'] = '8 (9644) 96-02-01'
            ws2['X103'] = 'Данилов Степан Юрьевич'
        elif city_word == 'Муром':
            shapka = 'ИП Трофимова Т.П. - TechPort.Murom'
            ws2['AK56'] = f'  {shapka}'
            ws2['BC60'] = '8 (903) 833-55-15'
        elif city_word == 'Ковров':
            shapka = 'ИП Трофимова Т.П. - TechPort.Kovrov'
            ws2['AK56'] = f'  {shapka}'
            ws2['BC60'] = '8 (920) 949-04-30'

        ws2['AD61'] = 'ПИПИПИПИПИПИПИПИ, именуемый в дальнейшем ' \
                      'Продавец, и _____________________________________________________, ' \
                      'именуемый в дальнейшем Покупатель, заключили настоящий  договор'

        if city_word == 'Владимир':
            p = ws2['AD61'].value
            px = p.replace('ПИПИПИПИПИПИПИПИ', 'ИП Трофимова Т.П. - TechPort.Vladimir')
            ws2['AD61'] = px
        elif city_word == 'Иваново':
            p = ws2['AD61'].value
            px = p.replace('ПИПИПИПИПИПИПИПИ', 'ИП Данилов С.Ю. - TechPort.Ivanovo')
            ws2['AD61'] = px
        elif city_word == 'Муром':
            p = ws2['AD61'].value
            px = p.replace('ПИПИПИПИПИПИПИПИ', 'ИП Трофимова Т.П. - TechPort.Murom')
            ws2['AD61'] = px
        elif city_word == 'Ковров':
            p = ws2['AD61'].value
            px = p.replace('ПИПИПИПИПИПИПИПИ', 'ИП Трофимова Т.П. - TechPort.Kovrov')
            ws2['AD61'] = px

        ws2.merge_cells(start_row=60, start_column=55, end_row=60, end_column=56)

        ws2['BC56'] = 'Отдел продаж'
        ws2.merge_cells(start_row=56, start_column=55, end_row=56, end_column=56)

        ws2['BC57'] = '8 (800) 555-87-78'
        ws2.merge_cells(start_row=57, start_column=55, end_row=57, end_column=56)

        ws2['BC59'] = 'Служба доставки'
        ws2.merge_cells(start_row=59, start_column=55, end_row=59, end_column=56)

        ws2.merge_cells('AD61:BD61')
        currentCell = ws2.cell(row=61, column=30)
        currentCell.alignment = Alignment(horizontal='left', vertical='top')
        ws2['AD61'].font = fonts.font8

        img2 = Image('logo.jpg')
        ws2.add_image(img2, 'AD56')

        # Наполнение сраной второй части страницы
        ws2.merge_cells(start_row=56, start_column=9, end_row=57, end_column=20)
        currentCell = ws2.cell(row=56, column=9)
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        ws2['I56'].font = fonts.font1

        ws2['Z56'] = 'Отдел продаж'
        ws2.merge_cells(start_row=56, start_column=26, end_row=56, end_column=27)
        currentCell = ws2.cell(row=56, column=26)
        currentCell.alignment = Alignment(horizontal='right', vertical='center')
        ws2['Z56'].font = fonts.font2

        ws2['Z57'] = '8 (800) 555-87-78'
        ws2.merge_cells(start_row=57, start_column=26, end_row=57, end_column=27)
        currentCell = ws2.cell(row=57, column=26)
        currentCell.alignment = Alignment(horizontal='right', vertical='center')
        ws2['Z57'].font = fonts.font2

        ws2['Z59'] = 'Служба доставки'
        ws2.merge_cells(start_row=59, start_column=26, end_row=59, end_column=27)
        currentCell = ws2.cell(row=59, column=26)
        currentCell.alignment = Alignment(horizontal='right', vertical='center')
        ws2['Z59'].font = fonts.font2

        ws2.merge_cells(start_row=60, start_column=26, end_row=60, end_column=27)
        currentCell = ws2.cell(row=60, column=26)
        currentCell.alignment = Alignment(horizontal='right', vertical='center')
        ws2['Z60'].font = fonts.font2

        ws2.cell(row=58, column=6).value = ws1['H4'].value + f' г. {city_changed(combo.get())}'
        ws2.merge_cells(start_row=58, start_column=6, end_row=59, end_column=24)
        currentCell = ws2.cell(row=58, column=6)
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        ws2['F58'].font = fonts.font1

        for row in ws2.iter_cols(min_col=2, max_col=28, min_row=61, max_row=68):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='bottom')

        ws2['B61'] = '7. ИНФОРМАЦИЯ О ТОВАРЕ И АДРЕСЕ ДОСТАВКИ'
        ws2.merge_cells(start_row=61, start_column=2, end_row=61, end_column=28)
        ws2['B61'].font = fonts.font3

        ws2['B62'] = 'Адрес:'
        ws2.merge_cells(start_row=62, start_column=2, end_row=62, end_column=3)
        ws2['B62'].font = fonts.font4

        # Адрес доставки
        ws2.merge_cells('F62:AB62')
        ws2['F62'].font = fonts.font1
        borders.set_border(ws2, 'F62:AB62')
        if city_word == 'Владимир':
            ws2['F62'] = 'г. Владимир'
            if ws1['F8'].value == None:
                ws2['F62'] = 'г. Владимир'
                if ws1['F8'].value == None:
                    ws2['F62'] = 'г. Владимир'
                else:
                    ws2.cell(row=62, column=6).value = ws1['F8'].value
            else:
                if len(ws1['F8'].value) == 0:
                    ws2['F62'] = 'г. Владимир'
                else:
                    ws2.cell(row=62, column=6).value = ws1['F8'].value

        elif city_word == 'Иваново':
            ws2['F62'] = 'г. Иваново'
            if ws1['F8'].value == None:
                ws2['F62'] = 'г. Иваново'
                if ws1['F8'].value == None:
                    ws2['F62'] = 'г. Иваново'
                else:
                    ws2.cell(row=62, column=6).value = ws1['F8'].value
            else:
                if len(ws1['F8'].value) == 0:
                    ws2['F62'] = 'г. Иваново'
                else:
                    ws2.cell(row=62, column=6).value = ws1['F8'].value

        elif city_word == 'Муром':
            ws2['F62'] = 'г. Муром'
            if ws1['F8'].value == None:
                ws2['F62'] = 'г. Муром'
                if ws1['F8'].value == None:
                    ws2['F62'] = 'г. Муром'
                else:
                    ws2.cell(row=62, column=6).value = ws1['F8'].value
            else:
                if len(ws1['F8'].value) == 0:
                    ws2['F62'] = 'г. Муром'
                else:
                    ws2.cell(row=62, column=6).value = ws1['F8'].value

        elif city_word == 'Ковров':
            ws2['F62'] = 'г. Ковров'
            if ws1['F8'].value == None:
                ws2['F62'] = 'г. Ковров'
                if ws1['F8'].value == None:
                    ws2['F62'] = 'г. Ковров'
                else:
                    ws2.cell(row=62, column=6).value = ws1['F8'].value
            else:
                if len(ws1['F8'].value) == 0:
                    ws2['F62'] = 'г. Ковров'
                else:
                    ws2.cell(row=62, column=6).value = ws1['F8'].value

        ws2['B64'] = 'Заказчик:'
        ws2.merge_cells(start_row=64, start_column=2, end_row=64, end_column=4)
        ws2['B64'].font = fonts.font4

        # Имя заказчика
        ws2.cell(row=64, column=6).value = ws1['F11'].value
        ws2.merge_cells('F64:L64')
        ws2['F64'].font = fonts.font1
        borders.set_border(ws2, 'F64:L64')

        ws2['N63'] = 'Телефоны:'
        ws2.merge_cells(start_row=63, start_column=14, end_row=64, end_column=15)
        ws2['N63'].font = fonts.font2

        # Телефон заказчика
        ws2.cell(row=63, column=17).value = ws1['Q10'].value
        ws2.merge_cells(start_row=63, start_column=17, end_row=64, end_column=27)
        ws2['Q63'].font = fonts.font5
        borders.set_border(ws2, 'Q64:AA64')

        ws2['B66'] = 'Доставка производится:     дата:'
        ws2.merge_cells('B66:I66')
        ws2['B66'].font = fonts.font2

        # Дата доставки
        ws2.cell(row=66, column=11).value = ws1['L13'].value
        ws2.merge_cells('K66:Q66')
        ws2['K66'].font = fonts.font1
        borders.set_border(ws2, 'K66:Q66')

        ws2['S66'] = 'время:'
        ws2.merge_cells('S66:U66')
        ws2['S66'].font = fonts.font2

        ws2['W66'] = 'с 12 до 19'
        ws2.merge_cells('W66:AB66')
        ws2['W66'].font = fonts.font1
        borders.set_border(ws2, 'W66:AB66')

        ws2['B68'] = 'Примечания:'
        ws2.merge_cells('B68:E68')
        ws2['B68'].font = fonts.font2

        # Условия доставки
        ws2['F68'] = 'Самовывоз'
        ws2.merge_cells('F68:Z68')
        ws2['F68'].font = fonts.font1
        borders.set_border(ws2, 'F68:Z68')
        if ws1['F15'].value == None:
            if ws1['F15'].value == None:
                ws2['F68'] = 'Самовывоз'
            elif ws1['F15'].value == 'null':
                ws2['F68'] = 'Самовывоз'
            else:
                ws2.cell(row=68, column=6).value = ws1['F15'].value
        elif ws1['F15'].value == 'null':
            ws2['F68'] = 'Самовывоз'
        else:
            if len(ws1['F15'].value) == 0:
                ws2['F68'] = 'Самовывоз'
            elif ws1['F15'].value == 'null':
                ws2['F68'] = 'Самовывоз'
            else:
                ws2.cell(row=68, column=6).value = ws1['F15'].value

        # Обьединение ячеек в таблице заказа
        for k in range(69, 85):
            ws2.merge_cells(start_row=k + 1, start_column=3, end_row=k + 1, end_column=13)
            ws2.merge_cells(start_row=k + 1, start_column=15, end_row=k + 1, end_column=17)
            ws2.merge_cells(start_row=k + 1, start_column=18, end_row=k + 1, end_column=21)
            ws2.merge_cells(start_row=k + 1, start_column=22, end_row=k + 1, end_column=25)

        # Границы в таблице заказа
        for col_cells in ws2.iter_cols(min_col=27, max_col=27, min_row=70, max_row=85):
            for cell in col_cells:
                cell.border = Border(left=Side(border_style='medium', color='000000'))

        for col_cells in ws2.iter_cols(min_col=1, max_col=1, min_row=70, max_row=85):
            for cell in col_cells:
                cell.border = Border(right=Side(border_style='medium', color='000000'))

        for col_cells in ws2.iter_cols(min_col=2, max_col=26, min_row=69, max_row=69):
            for cell in col_cells:
                cell.border = Border(bottom=Side(border_style='medium', color='000000'))

        for col_cells in ws2.iter_cols(min_col=2, max_col=26, min_row=70, max_row=83):
            for cell in col_cells:
                cell.border = Border(bottom=Side(border_style='thin', color='000000'),
                                     right=Side(border_style='thin', color='000000'))

        for col_cells in ws2.iter_cols(min_col=2, max_col=26, min_row=84, max_row=84):
            for cell in col_cells:
                cell.border = Border(bottom=Side(border_style='medium', color='000000'),
                                     right=Side(border_style='thin', color='000000'))

        for col_cells in ws2.iter_cols(min_col=2, max_col=26, min_row=85, max_row=85):
            for cell in col_cells:
                cell.border = Border(bottom=Side(border_style='medium', color='000000'),
                                     right=Side(border_style='thin', color='000000'))

        # Выравнивание ячеек в таблице заказа
        for row in ws2.iter_cols(min_col=14, max_col=26, min_row=70, max_row=85):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        # пп 1 - 14
        for row in ws2.iter_cols(min_col=2, max_col=2, min_row=70, max_row=85):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws2.iter_cols(min_col=3, max_col=13, min_row=71, max_row=84):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for row in ws2.iter_cols(min_col=14, max_col=26, min_row=85, max_row=85):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Шрифт в таблице заказа
        for row in ws2.iter_cols(min_col=3, max_col=26, min_row=71, max_row=84):
            for cell in row:
                cell.font = fonts.font3

        for row in ws2.iter_cols(min_col=3, max_col=26, min_row=85, max_row=85):
            for cell in row:
                cell.font = fonts.font6

        # пп 1 - 14
        for row in ws2.iter_cols(min_col=2, max_col=2, min_row=70, max_row=84):
            for cell in row:
                cell.font = fonts.font6

        # наименование - примечание
        for row in ws2.iter_cols(min_col=3, max_col=26, min_row=70, max_row=70):
            for cell in row:
                cell.font = fonts.font6

        ws2['B70'] = 'пп'
        ws2['B71'] = int('1')
        ws2['B72'] = int('2')
        ws2['B73'] = int('3')
        ws2['B74'] = int('4')
        ws2['B75'] = int('5')
        ws2['B76'] = int('6')
        ws2['B77'] = int('7')
        ws2['B78'] = int('8')
        ws2['B79'] = int('9')
        ws2['B80'] = int('10')
        ws2['B81'] = int('11')
        ws2['B82'] = int('12')
        ws2['B83'] = int('13')
        ws2['B84'] = int('14')

        ws2['C70'] = 'наименование'
        currentCell = ws2.cell(row=70, column=3)
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws2['N70'] = 'к-во'
        ws2['O70'] = 'мест'
        ws2['R70'] = 'цена'
        ws2['V70'] = 'сумма'
        ws2['Z70'] = 'примечание'
        ws2['C85'] = 'итого:'
        currentCell = ws2.cell(row=85, column=3)
        currentCell.alignment = Alignment(horizontal='right', vertical='center')

        # Первая позиция
        def pp1_1():
            for i in ws2:
                if ws1['B18'].value == 1:
                    # Наименование
                    ws2.cell(row=71, column=3).value = ws1['C18'].value
                    # К-во
                    ws2.cell(row=71, column=14).value = ws1['N18'].value
                    # мест
                    ws2.cell(row=71, column=15).value = ws1['O18'].value
                    # цена
                    ws2.cell(row=71, column=18).value = ws1['R18'].value
                    # сумма
                    ws2.cell(row=71, column=22).value = ws1['V18'].value
                    # примечание
                    ws2.cell(row=71, column=26).value = ws1['Z18'].value

                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C71'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R71'] = c
                            ws2['V71'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R71'] = c
                            ws2['V71'] = c
                    else:
                        break

        pp1_1()

        # Вторая позиция
        def pp2_2():
            for n in ws2:
                if ws1['B19'].value == 2:
                    # Наименование
                    ws2.cell(row=72, column=3).value = ws1['C19'].value
                    # К-во
                    ws2.cell(row=72, column=14).value = ws1['N19'].value
                    # мест
                    ws2.cell(row=72, column=15).value = ws1['O19'].value
                    # цена
                    ws2.cell(row=72, column=18).value = ws1['R19'].value
                    # сумма
                    ws2.cell(row=72, column=22).value = ws1['V19'].value
                    # примечание
                    ws2.cell(row=72, column=26).value = ws1['Z19'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C72'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R72'] = c
                            ws2['V72'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R72'] = c
                            ws2['V72'] = c
                    else:
                        break

        if ws1['C18'].value == None:
            ws2['C72'] = ''
        elif ws2['C71'].value == ws1['C18'].value:
            pp2_2()

        # Третья позиция
        def pp3_3():
            for n in ws2:
                if ws1['B20'].value == 3:
                    # Наименование
                    ws2.cell(row=73, column=3).value = ws1['C20'].value
                    # К-во
                    ws2.cell(row=73, column=14).value = ws1['N20'].value
                    # мест
                    ws2.cell(row=73, column=15).value = ws1['O20'].value
                    # цена
                    ws2.cell(row=73, column=18).value = ws1['R20'].value
                    # сумма
                    ws2.cell(row=73, column=22).value = ws1['V20'].value
                    # примечание
                    ws2.cell(row=73, column=26).value = ws1['Z20'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C73'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R73'] = c
                            ws2['V73'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R73'] = c
                            ws2['V73'] = c
                    else:
                        break

        if ws1['C19'].value == None:
            ws2['C73'] = ''
        elif ws2['C72'].value == ws1['C19'].value:
            pp3_3()

        # Четвертая позиция
        def pp4_4():
            for n in ws2:
                if ws1['B21'].value == 4:
                    # Наименование
                    ws2.cell(row=74, column=3).value = ws1['C21'].value
                    # К-воow
                    ws2.cell(row=74, column=14).value = ws1['N21'].value
                    # мест
                    ws2.cell(row=74, column=15).value = ws1['O21'].value
                    # цена
                    ws2.cell(row=74, column=18).value = ws1['R21'].value
                    # сумма
                    ws2.cell(row=74, column=22).value = ws1['V21'].value
                    # примечание
                    ws2.cell(row=74, column=26).value = ws1['Z21'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C74'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R74'] = c
                            ws2['V74'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R74'] = c
                            ws2['V74'] = c
                    else:
                        break

        if ws1['C20'].value == None:
            ws2['C74'] = ''
        elif ws2['C73'].value == ws1['C20'].value:
            pp4_4()

        # Пятая позиция
        def pp5_5():
            for n in ws2:
                if ws1['B22'].value == 5:
                    # Наименование
                    ws2.cell(row=75, column=3).value = ws1['C22'].value
                    # К-во
                    ws2.cell(row=75, column=14).value = ws1['N22'].value
                    # мест
                    ws2.cell(row=75, column=15).value = ws1['O22'].value
                    # цена
                    ws2.cell(row=75, column=18).value = ws1['R22'].value
                    # сумма
                    ws2.cell(row=75, column=22).value = ws1['V22'].value
                    # примечание
                    ws2.cell(row=75, column=26).value = ws1['Z22'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C75'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R75'] = c
                            ws2['V75'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R75'] = c
                            ws2['V75'] = c
                    else:
                        break

        if ws1['C21'].value == None:
            ws2['C75'] = ''
        elif ws2['C74'].value == ws1['C21'].value:
            pp5_5()

        # Шестая позиция
        def pp6_6():
            for n in ws2:
                if ws1['B23'].value == 6:
                    # Наименование
                    ws2.cell(row=76, column=3).value = ws1['C23'].value
                    # К-во
                    ws2.cell(row=76, column=14).value = ws1['N23'].value
                    # мест
                    ws2.cell(row=76, column=15).value = ws1['O23'].value
                    # цена
                    ws2.cell(row=76, column=18).value = ws1['R23'].value
                    # сумма
                    ws2.cell(row=76, column=22).value = ws1['V23'].value
                    # примечание
                    ws2.cell(row=76, column=26).value = ws1['Z23'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C76'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R76'] = c
                            ws2['V76'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R76'] = c
                            ws2['V76'] = c
                    else:
                        break

        if ws1['C22'].value == None:
            ws2['C76'] = ''
        elif ws2['C75'].value == ws1['C22'].value:
            pp6_6()

        # Седьмая позиция
        def pp7_7():
            for n in ws2:
                if ws1['B24'].value == 7:
                    # Наименование
                    ws2.cell(row=77, column=3).value = ws1['C24'].value
                    # К-во
                    ws2.cell(row=77, column=14).value = ws1['N24'].value
                    # мест
                    ws2.cell(row=77, column=15).value = ws1['O24'].value
                    # цена
                    ws2.cell(row=77, column=18).value = ws1['R24'].value
                    # сумма
                    ws2.cell(row=77, column=22).value = ws1['V24'].value
                    # примечание
                    ws2.cell(row=77, column=26).value = ws1['Z24'].value
                else:

                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C77'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R77'] = c
                            ws2['V77'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R77'] = c
                            ws2['V77'] = c
                    else:
                        break

        if ws1['C23'].value == None:
            ws2['C77'] = ''
        elif ws2['C76'].value == ws1['C23'].value:
            pp7_7()

        # Восьмая позиция
        def pp8_8():
            for n in ws2:
                if ws1['B25'].value == 8:
                    # Наименование
                    ws2.cell(row=78, column=3).value = ws1['C25'].value
                    # К-во
                    ws2.cell(row=78, column=14).value = ws1['N25'].value
                    # мест
                    ws2.cell(row=78, column=15).value = ws1['O25'].value
                    # цена
                    ws2.cell(row=78, column=18).value = ws1['R25'].value
                    # сумма
                    ws2.cell(row=78, column=22).value = ws1['V25'].value
                    # примечание
                    ws2.cell(row=78, column=26).value = ws1['Z25'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C78'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R78'] = c
                            ws2['V78'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R78'] = c
                            ws2['V78'] = c
                    else:
                        break

        if ws1['C24'].value == None:
            ws2['C278'] = ''
        elif ws2['C77'].value == ws1['C24'].value:
            pp8_8()

        # Девятая позиция
        def pp9_9():
            for n in ws2:
                if ws1['B26'].value == 9:
                    # Наименование
                    ws2.cell(row=79, column=3).value = ws1['C26'].value
                    # К-во
                    ws2.cell(row=79, column=14).value = ws1['N26'].value
                    # мест
                    ws2.cell(row=79, column=15).value = ws1['O26'].value
                    # цена
                    ws2.cell(row=79, column=18).value = ws1['R26'].value
                    # сумма
                    ws2.cell(row=79, column=22).value = ws1['V26'].value
                    # примечание
                    ws2.cell(row=79, column=26).value = ws1['Z26'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C79'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R79'] = c
                            ws2['V79'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R79'] = c
                            ws2['V79'] = c
                    else:
                        break

        if ws1['C25'].value == None:
            ws2['C79'] = ''
        elif ws2['C78'].value == ws1['C25'].value:
            pp9_9()

        # Десятая позиция
        def pp10_10():
            for n in ws2:
                if ws1['B27'].value == 10:
                    # Наименование
                    ws2.cell(row=80, column=3).value = ws1['C27'].value
                    # К-во
                    ws2.cell(row=80, column=14).value = ws1['N27'].value
                    # мест
                    ws2.cell(row=80, column=15).value = ws1['O27'].value
                    # цена
                    ws2.cell(row=80, column=18).value = ws1['R27'].value
                    # сумма
                    ws2.cell(row=80, column=22).value = ws1['V27'].value
                    # примечание
                    ws2.cell(row=80, column=26).value = ws1['Z27'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C80'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R80'] = c
                            ws2['V80'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R80'] = c
                            ws2['V80'] = c
                    else:
                        break

        if ws1['C26'].value == None:
            ws2['C80'] = ''
        elif ws2['C79'].value == ws1['C26'].value:
            pp10_10()

        # Одиннадцатая позиция
        def pp11_11():
            for n in ws2:
                if ws1['B28'].value == 11:
                    # Наименование
                    ws2.cell(row=81, column=3).value = ws1['C28'].value
                    # К-во
                    ws2.cell(row=81, column=14).value = ws1['N28'].value
                    # мест
                    ws2.cell(row=81, column=15).value = ws1['O28'].value
                    # цена
                    ws2.cell(row=81, column=18).value = ws1['R28'].value
                    # сумма
                    ws2.cell(row=81, column=22).value = ws1['V28'].value
                    # примечание
                    ws2.cell(row=81, column=26).value = ws1['Z28'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C81'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R81'] = c
                            ws2['V81'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R81'] = c
                            ws2['V81'] = c
                    else:
                        break

        if ws1['C27'].value == None:
            ws2['C81'] = ''
        elif ws2['C80'].value == ws1['C27'].value:
            pp11_11()

        # Двеннадцатая позиция
        def pp12_12():
            for n in ws2:
                if ws1['B29'].value == 12:
                    # Наименование
                    ws2.cell(row=82, column=3).value = ws1['C29'].value
                    # К-во
                    ws2.cell(row=82, column=14).value = ws1['N29'].value
                    # мест
                    ws2.cell(row=82, column=15).value = ws1['O29'].value
                    # цена
                    ws2.cell(row=82, column=18).value = ws1['R29'].value
                    # сумма
                    ws2.cell(row=82, column=22).value = ws1['V29'].value
                    # примечание
                    ws2.cell(row=82, column=26).value = ws1['Z29'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C82'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R82'] = c
                            ws2['V82'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R82'] = c
                            ws2['V82'] = c
                    else:
                        break

        if ws1['C28'].value == None:
            ws2['C82'] = ''
        elif ws2['C81'].value == ws1['C28'].value:
            pp12_12()

        # Тринадцатая позиция
        def pp13_13():
            for n in ws2:
                if ws1['B30'].value == 13:
                    # Наименование
                    ws2.cell(row=83, column=3).value = ws1['C30'].value
                    # К-во
                    ws2.cell(row=83, column=14).value = ws1['N30'].value
                    # мест
                    ws2.cell(row=83, column=15).value = ws1['O30'].value
                    # цена
                    ws2.cell(row=83, column=18).value = ws1['R30'].value
                    # сумма
                    ws2.cell(row=83, column=22).value = ws1['V30'].value
                    # примечание
                    ws2.cell(row=83, column=26).value = ws1['Z30'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C83'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R83'] = c
                            ws2['V83'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R83'] = c
                            ws2['V83'] = c
                    else:
                        break

        if ws1['C29'].value == None:
            ws2['C83'] = ''
        elif ws2['C82'].value == ws1['C29'].value:
            pp13_13()

        # Четырнадцатая позиция
        def pp14_14():
            for n in ws2:
                if ws1['B31'].value == 14:
                    # Наименование
                    ws2.cell(row=84, column=3).value = ws1['C31'].value
                    # К-во
                    ws2.cell(row=84, column=14).value = ws1['N31'].value
                    # мест
                    ws2.cell(row=84, column=15).value = ws1['O31'].value
                    # цена
                    ws2.cell(row=84, column=18).value = ws1['R31'].value
                    # сумма
                    ws2.cell(row=84, column=22).value = ws1['V31'].value
                    # примечание
                    ws2.cell(row=84, column=26).value = ws1['Z31'].value
                else:
                    if 'дост' in ws2['F14'].value:
                        mystring = ws2['F14'].value
                        ws2['C84'] = 'Доставка'
                        try:
                            result = re.findall(r'дост\s\d\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R84'] = c
                            ws2['V84'] = c
                        except:
                            result = re.findall(r'дост\s\d\d\w', mystring)
                            q = ''.join(result)
                            keyword = 'дост'
                            before_keyword, keyword, after_keyword = q.partition(keyword)
                            c = int(after_keyword)
                            ws2['R84'] = c
                            ws2['V84'] = c
                    else:
                        break

        if ws1['C29'].value == None:
            ws2['C84'] = ''
        elif ws2['C83'].value == ws1['C30'].value:
            pp14_14()

        ws2['N85'] = '=SUM(N71:N84)'
        ws2['O85'] = '=SUM(O71:O84)'
        ws2['V85'] = '=SUM(V71:V84)'

        ws2['B87'] = 'АКТ ПРИЕМА-ПЕРЕДАЧИ '
        ws2.merge_cells('B87:Z87')
        currentCell = ws2.cell(row=87, column=2)
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        ws2['B87'].font = fonts.font6

        ws2.cell(row=88, column=2).value = ws2['B34'].value
        ws2.merge_cells(start_row=88, start_column=2, end_row=97, end_column=26)
        currentCell = ws2.cell(row=88, column=2)
        currentCell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws2['B88'].font = fonts.font7

        for row in ws2.iter_cols(min_col=2, max_col=26, min_row=98, max_row=102):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)

        ws2['B98'] = '8. ПОДПИСИ СТОРОН'
        ws2.merge_cells('B98:Z98')
        ws2['B98'].font = fonts.font3

        ws2['B99'] = 'Настоящий Договор составлен в двух экземплярах, первый остается ' \
                     'у представителя Продавца, второй экземпляр выдается Покупателю. ' \
                     'Договор вступает в силу с момента его подписания сторонами и действует ' \
                     'до момента надлежащего исполнения сторонами принятых на себя обязательств.'
        ws2.merge_cells(start_row=99, start_column=2, end_row=101, end_column=26)
        ws2['B99'].font = fonts.font7

        ws2['B102'] = 'С условиями настоящего договора ознакомлен до момента его подписания и полностью согласен'
        ws2.merge_cells('B102:Z102')
        ws2['B102'].font = fonts.font7

        ws2['K103'] = 'Представитель Продавца'
        ws2.merge_cells('K103:M103')
        currentCell = ws2.cell(row=103, column=11)
        currentCell.alignment = Alignment(horizontal='right', vertical='bottom')
        ws2['K103'].font = fonts.font3

        borders.set_border(ws2, 'N103:U103')

        for row in ws2.iter_cols(min_col=14, max_col=26, min_row=103, max_row=104):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='bottom')

        ws2.merge_cells('X103:Z103')
        ws2['X103'].font = fonts.font3
        borders.set_border(ws2, 'X103:Z103')

        ws2['N104'] = '(Подпись)'
        ws2.merge_cells('N104:V104')
        ws2['N104'].font = fonts.font7

        ws2['X104'] = '(ф.и.о.)'
        ws2.merge_cells('X104:Z104')
        ws2['X104'].font = fonts.font7

        ws2['B106'] = 'Покупатель'
        ws2.merge_cells('B52:D52')
        currentCell = ws2.cell(row=106, column=2)
        currentCell.alignment = Alignment(horizontal='left', vertical='bottom')
        ws2['B106'].font = fonts.font1

        borders.set_border(ws2, 'E106:H106')
        borders.set_border(ws2, 'K106:Q106')

        ws2['R106'] = 'Дата '
        ws2.merge_cells('R106:T106')
        currentCell = ws2.cell(row=106, column=18)
        currentCell.alignment = Alignment(horizontal='right', vertical='bottom')
        ws2['R106'].font = fonts.font1

        ws2['U106'] = '«       »                                             2022г.'
        ws2.merge_cells('U106:Z106')
        currentCell = ws2.cell(row=106, column=21)
        currentCell.alignment = Alignment(horizontal='left', vertical='bottom')
        ws2['U106'].font = fonts.font1
        borders.set_border(ws2, 'U106:Z106')

        ws2['E107'] = ' (Подпись)'
        ws2.merge_cells('E107:H107')
        currentCell = ws2.cell(row=107, column=5)
        currentCell.alignment = Alignment(horizontal='center', vertical='bottom')
        ws2['E107'].font = fonts.font6

        ws2['K107'] = ' (ф.и.о.)'
        ws2.merge_cells('K107:Q107')
        currentCell = ws2.cell(row=107, column=11)
        currentCell.alignment = Alignment(horizontal='center', vertical='bottom')
        ws2['K107'].font = fonts.font6

        img2 = Image('logo.jpg')
        ws2.add_image(img2, 'B56')

        def save():
            date_format = '%d.%m.%Y'
            today = datetime.now().date()
            tomorrow = today + timedelta(days=1)
            dd = tomorrow.strftime(date_format)

            number = ws2['F4'].value
            res = re.findall(r'\w+', number)
            word = res[4]

            pyt = lbl.cget('text')
            wb2.save(f'{pyt}/{dd} {word}.xlsx')

        save()

        '''window.update_idletasks()
        bar['value'] += i
        time.sleep(.3)
        txt['text'] = bar['value'], '%' '''

# Кнопка
btn = Button(tab1,
             text='Добавить файлы',
             height=2,
             width=13,
             command=clicked,
             font=('Arial Bold', 10),
             bg='steelblue3',
             fg='white')
btn.grid(column=2, row=3, pady=5, sticky=N)

btn5 = Button(tab1,
              text='Выбрать все',
              height=2,
              width=13,
              command=select,
              font=('Arial Bold', 10),
              bg='steelblue3',
              fg='white')
btn5.grid(column=2, row=6, sticky=N)

btn2 = Button(tab1,
              text='Удалить все',
              height=2,
              width=13,
              command=delete_all,
              font=('Arial Bold', 10))
btn2.grid(column=2, row=4, sticky=N)

btn3 = Button(tab1,
              text='Удалить',
              height=2,
              width=13,
              command=delete,
              font=('Arial Bold', 10))
btn3.grid(column=2, row=5, sticky=N)

btn4 = Button(tab1,
              text='Набрать!',
              height=2,
              width=13,
              command=nabor,
              font=('Arial Bold', 10),
              bg='steelblue3',
              fg='white')
btn4.grid(column=2, row=7, sticky=N)

# ListBox, Label, Progressbar
lb_p = Listbox(tab2, height=2, width=60, selectmode=EXTENDED)
lb_p.grid(column=0, row=3, columnspan=2, padx=3)


def clicked_p():
    file_open_p = filedialog.askopenfilenames()
    names = file_open_p
    for pp in names:
        lb_p.insert(END, pp)


lbl_p = Label(tab2,
            text='Нажмите кнопку и выберите папку для сохранения            ',
            font=('Arial Bold', 10),
            bg='white',
            height=2,
            width=44)
lbl_p.grid(column=0, row=4, columnspan=2, padx=3, pady=8)


def savefold_p():
    dire_p = filedialog.askdirectory()
    name = dire_p
    for pp in name:
        lbl_p.configure(text=f'{name}')


def nabor_p():
    r = lb_p.get(0, 70)
    for i in range(0, len(r), 1):
        file = r[i:i + 1]
        file_p = ''.join(file)
        with warnings.catch_warnings(record=True):
            warnings.simplefilter("always")
            wb = openpyxl.load_workbook(file_p)
            ws = wb.active

            row_max = ws.max_row

            for j in range(7):
                ws.delete_cols(10)

            for i in range(2, row_max):
                # Устанавливаю стандартные наценки по цене
                if int(ws[f'H{i}'].value) < 200:
                    ws[f'I{i}'] = f'=CEILING(IF(G{i}>((H{i})+20),G{i},H{i}+20),10)'
                elif int(ws[f'H{i}'].value) > 200 and int(ws[f'H{i}'].value) < 2000:
                    ws[f'I{i}'] = f'=CEILING(IF(G{i}>((H{i})*1.1),G{i},H{i}*1.1),10)'
                elif int(ws[f'H{i}'].value) > 2000 and int(ws[f'H{i}'].value) < 5000:
                    ws[f'I{i}'] = f'=CEILING(IF(G{i}>((H{i})+250),G{i},H{i}+250),10)'
                elif int(ws[f'H{i}'].value) > 5000 and int(ws[f'H{i}'].value) < 20000:
                    ws[f'I{i}'] = f'=CEILING(IF(G{i}>((H{i})+600),G{i},H{i}+600),10)'
                elif int(ws[f'H{i}'].value) > 20000 and int(ws[f'H{i}'].value) < 30000:
                    ws[f'I{i}'] = f'=CEILING(IF(G{i}>((H{i})+1100),G{i},H{i}+1100),10)'
                elif int(ws[f'H{i}'].value) > 30000 and int(ws[f'H{i}'].value) < 40000:
                    ws[f'I{i}'] = f'=CEILING(IF(G{i}>((H{i})+1600),G{i},H{i}+1600),10)'
                elif int(ws[f'H{i}'].value) > 40000 and int(ws[f'H{i}'].value) < 50000:
                    ws[f'I{i}'] = f'=CEILING(IF(G{i}>((H{i})+2100),G{i},H{i}+2100),10)'
                elif int(ws[f'H{i}'].value) > 50000 and int(ws[f'H{i}'].value) < 70000:
                    ws[f'I{i}'] = f'=CEILING(IF(G{i}>((H{i})+2600),G{i},H{i}+2600),10)'
                elif int(ws[f'H{i}'].value) > 70000:
                    ws[f'I{i}'] = f'=CEILING(IF(G{i}>((H{i})+3100),G{i},H{i}+3100),10)'

                # Мебель
                if 'мебель' in ws[f'C{i}'].value:
                    ws[f'I{i}'] = f'=CEILING(IF(G{i}>((H{i})*1.1),G{i},H{i}*1.1),10)'
                elif 'Мебель' in ws[f'C{i}'].value:
                    ws[f'I{i}'] = f'=CEILING(IF(G{i}>((H{i})*1.1),G{i},H{i}*1.1),10)'
                    # Зоотовары
                elif 'Зоотовары' in ws[f'C{i}'].value:
                    ws[f'I{i}'] = f'=CEILING(IF(G{i}>((H{i})*1.13),G{i},H{i}*1.13),10)'
                    # БХИГ
                elif 'БХИГ' in ws[f'C{i}'].value:
                    ws[f'I{i}'] = f'=CEILING(IF(G{i}>((H{i})*1.15),G{i},H{i}*1.15),10)'
                    # Химия
                elif 'Химия' in ws[f'C{i}'].value:
                    ws[f'I{i}'] = f'=CEILING(IF(G{i}>((H{i})*1.15),G{i},H{i}*1.15),10)'
                    # Отделка
                elif 'Отделка' in ws[f'C{i}'].value:
                    ws[f'I{i}'] = f'=CEILING(IF(G{i}>((H{i})*1.2),G{i},H{i}*1.2),10)'
                    # Септики и системы очистки
                elif 'Септики и системы очистки' in ws[f'C{i}'].value:
                    ws[f'I{i}'] = f'=CEILING(IF(G{i}>((H{i})*1.3),G{i},H{i}*1.3),10)'
                    # ЭкоПром
                elif 'ЭкоПром' in ws[f'D{i}'].value:
                    ws[f'I{i}'] = f'=CEILING(IF(G{i}>((H{i})*1.3),G{i},H{i}*1.3),10)'


                lbl_p2.configure(text='Готово!', font=('Arial Bold', 10), bg='green', fg='white')


            def save_p():
                date_format = '%d.%m.%Y'
                today = datetime.now().date()
                dd1 = today.strftime(date_format)

                pyt = lbl_p.cget('text')
                wb.save(f'{pyt}/Прайс Владимир на {dd1}.xlsx')

            save_p()

lbl_p2 = Label(tab2,
            text='Процесс выполнения набора',
            font=('Arial Bold', 10),
            bg='white',
            height=2,
            width=44)
lbl_p2.grid(column=0, row=6, columnspan=2, padx=3, pady=8)


def select_p():
    lb_p.select_set(0, END)


def delete_all_p():
    lb_p.delete(0, 'end')


# Кнопки для окна "Прайс"
btn_p = Button(tab2,
             text='Добавить прайс',
             height=2,
             width=13,
             command=clicked_p,
             font=('Arial Bold', 10),
             bg='steelblue3',
             fg='white')
btn_p.grid(column=2, row=3, pady=8, sticky=N)

btn_s = Button(tab2,
             text='Сохранить в...',
             height=2,
             width=13,
             command=savefold_p,
             font=('Arial Bold', 10),
             bg='steelblue3',
             fg='white')
btn_s.grid(column=2, row=4, pady=8, sticky=N)

btn_d = Button(tab2,
              text='Удалить все',
              height=2,
              width=13,
              command=delete_all_p,
              font=('Arial Bold', 10))
btn_d.grid(column=2, row=5, sticky=N)

btn_s = Button(tab2,
             text='Выбрать все',
             height=2,
             width=13,
             command=select_p,
             font=('Arial Bold', 10),
             bg='steelblue3',
             fg='white')
btn_s.grid(column=2, row=6, pady=5, sticky=N)

btn_n = Button(tab2,
             text='Набрать прайс!',
             height=2,
             width=13,
             command=nabor_p,
             font=('Arial Bold', 10),
             bg='steelblue3',
             fg='white')
btn_n.grid(column=2, row=7, pady=5, sticky=N)

window.mainloop()